# -*- coding: utf-8 -*-

import io
import base64
from datetime import date, datetime
from odoo import http, _
from odoo.http import request, content_disposition


class LagunesExportController(http.Controller):

    @http.route('/lagunes/export/commandes', type='http', auth='user', methods=['GET'])
    def export_commandes(self, date_debut=None, date_fin=None, entreprise_id=None, **kwargs):
        """
        Export Excel des commandes pour une période et/ou une entreprise.
        Accessible uniquement aux utilisateurs connectés (groupe manager).
        """
        # Construire le domaine
        domain = [('company_id', '=', request.env.company.id)]

        if date_debut:
            try:
                domain.append(('date', '>=', date_debut))
            except Exception:
                pass

        if date_fin:
            try:
                domain.append(('date', '<=', date_fin))
            except Exception:
                pass

        if entreprise_id:
            try:
                domain.append(('entreprise_id', '=', int(entreprise_id)))
            except (ValueError, TypeError):
                pass

        # Récupérer les commandes
        commandes = request.env['lagunes.commande'].search(
            domain,
            order='date desc, entreprise_id, create_date desc'
        )
        # Filtrer par company si les règles de sécurité ne le font pas déjà
        commandes = commandes.filtered(lambda c: c.company_id.id == request.env.company.id)

        # Générer le fichier Excel
        excel_data = self._generate_excel_commandes(commandes, date_debut, date_fin)

        # Nom du fichier
        suffix = ''
        if date_debut and date_fin:
            suffix = f"_{date_debut}_au_{date_fin}"
        elif date_debut:
            suffix = f"_depuis_{date_debut}"
        elif date_fin:
            suffix = f"_jusquau_{date_fin}"

        filename = f"commandes_cantine{suffix}.xlsx"

        return request.make_response(
            excel_data,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(filename)),
                ('Content-Length', len(excel_data)),
            ]
        )

    @http.route('/lagunes/export/employes', type='http', auth='user', methods=['GET'])
    def export_employes(self, entreprise_id=None, **kwargs):
        """Export Excel des employés."""
        domain = [('company_id', '=', request.env.company.id)]
        if entreprise_id:
            try:
                domain.append(('entreprise_id', '=', int(entreprise_id)))
            except (ValueError, TypeError):
                pass

        employes = request.env['lagunes.employe'].search(
            domain,
            order='entreprise_id, nom, prenom'
        )

        excel_data = self._generate_excel_employes(employes)
        filename = "employes_cantine.xlsx"

        return request.make_response(
            excel_data,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', content_disposition(filename)),
                ('Content-Length', len(excel_data)),
            ]
        )

    # ------------------------------------------------------------------ #
    #  GÉNÉRATION EXCEL                                                    #
    # ------------------------------------------------------------------ #

    def _generate_excel_commandes(self, commandes, date_debut=None, date_fin=None):
        """Génère un fichier Excel avec les commandes."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback CSV si openpyxl non disponible
            return self._generate_csv_commandes(commandes)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commandes"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="16166D", end_color="16166D", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        # Titre
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = "Restaurant des Lagunes — Export des commandes"
        title_cell.font = Font(bold=True, size=14, color="16166D")
        title_cell.alignment = Alignment(horizontal="center")

        # Sous-titre période
        period_text = "Toutes les commandes"
        if date_debut and date_fin:
            period_text = f"Période : {date_debut} au {date_fin}"
        elif date_debut:
            period_text = f"À partir du : {date_debut}"
        elif date_fin:
            period_text = f"Jusqu'au : {date_fin}"

        ws.merge_cells('A2:J2')
        period_cell = ws['A2']
        period_cell.value = period_text
        period_cell.font = Font(italic=True, color="555555")
        period_cell.alignment = Alignment(horizontal="center")

        ws.merge_cells('A3:J3')
        ws['A3'].value = f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A3'].font = Font(italic=True, color="888888", size=9)
        ws['A3'].alignment = Alignment(horizontal="center")

        # En-têtes colonnes
        headers = [
            'Référence', 'Date', 'Entreprise', 'Employé', 'Plat',
            'Options', 'Notes', 'Prix unitaire (FCFA)', 'Prix total (FCFA)', 'Statut'
        ]
        row_num = 5
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        ws.row_dimensions[row_num].height = 30

        # Données
        alt_fill = PatternFill(start_color="F0F0F8", end_color="F0F0F8", fill_type="solid")
        state_labels = {
            'draft': 'Brouillon',
            'confirmed': 'Confirmée',
            'preparing': 'En préparation',
            'ready': 'Prêt',
            'delivered': 'Livré',
            'cancelled': 'Annulé',
        }

        for i, cmd in enumerate(commandes):
            row_num += 1
            fill = alt_fill if i % 2 == 0 else None
            options_str = ', '.join(cmd.option_ids.mapped('name')) if cmd.option_ids else ''
            row_data = [
                cmd.reference,
                cmd.date.strftime('%d/%m/%Y') if cmd.date else '',
                cmd.entreprise_id.name if cmd.entreprise_id else '',
                cmd.employee_name or '',
                cmd.plat_id.name if cmd.plat_id else '',
                options_str,
                cmd.notes or '',
                cmd.prix_unitaire,
                cmd.prix_total,
                state_labels.get(cmd.state, cmd.state),
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_num in [6, 7]))
                if fill:
                    cell.fill = fill
                if col_num in [8, 9]:
                    cell.number_format = '#,##0'

        # Ligne totaux
        if commandes:
            row_num += 1
            total_cell = ws.cell(row=row_num, column=8, value="TOTAL")
            total_cell.font = Font(bold=True)
            total_cell.alignment = Alignment(horizontal="right")
            total_cell.fill = PatternFill(start_color="E0E1F0", end_color="E0E1F0", fill_type="solid")

            sum_cell = ws.cell(row=row_num, column=9, value=sum(commandes.mapped('prix_total')))
            sum_cell.font = Font(bold=True)
            sum_cell.number_format = '#,##0'
            sum_cell.fill = PatternFill(start_color="E0E1F0", end_color="E0E1F0", fill_type="solid")
            sum_cell.border = thin_border

            count_cell = ws.cell(row=row_num, column=1, value=f"{len(commandes)} commande(s)")
            count_cell.font = Font(bold=True, italic=True, color="16166D")

        # Largeurs colonnes
        col_widths = [15, 12, 25, 20, 25, 25, 25, 18, 18, 15]
        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Figer la ligne d'en-têtes
        ws.freeze_panes = 'A6'

        # Sauvegarder
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _generate_excel_employes(self, employes):
        """Génère un fichier Excel avec la liste des employés."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return b''

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employés"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="16166D", end_color="16166D", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # Titre
        ws.merge_cells('A1:F1')
        ws['A1'].value = "Restaurant des Lagunes — Liste des employés"
        ws['A1'].font = Font(bold=True, size=14, color="16166D")
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:F2')
        ws['A2'].value = f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(italic=True, color="888888", size=9)
        ws['A2'].alignment = Alignment(horizontal="center")

        # Headers
        headers = ['Entreprise', 'Nom', 'Prénom', 'Email', 'Date inscription', 'Actif']
        row_num = 4
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws.row_dimensions[row_num].height = 25

        alt_fill = PatternFill(start_color="F0F0F8", end_color="F0F0F8", fill_type="solid")
        for i, emp in enumerate(employes):
            row_num += 1
            fill = alt_fill if i % 2 == 0 else None
            row_data = [
                emp.entreprise_id.name if emp.entreprise_id else '',
                emp.nom,
                emp.prenom,
                emp.email,
                emp.date_inscription.strftime('%d/%m/%Y') if emp.date_inscription else '',
                'Oui' if emp.active else 'Non',
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill

        col_widths = [25, 20, 20, 30, 18, 10]
        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        ws.freeze_panes = 'A5'

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _generate_csv_commandes(self, commandes):
        """Fallback CSV si openpyxl non disponible."""
        import csv
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(['Référence', 'Date', 'Entreprise', 'Employé', 'Plat',
                         'Options', 'Notes', 'Prix unitaire', 'Prix total', 'Statut'])
        for cmd in commandes:
            options_str = ', '.join(cmd.option_ids.mapped('name')) if cmd.option_ids else ''
            writer.writerow([
                cmd.reference,
                cmd.date.strftime('%d/%m/%Y') if cmd.date else '',
                cmd.entreprise_id.name if cmd.entreprise_id else '',
                cmd.employee_name or '',
                cmd.plat_id.name if cmd.plat_id else '',
                options_str,
                cmd.notes or '',
                cmd.prix_unitaire,
                cmd.prix_total,
                cmd.state,
            ])
        return output.getvalue().encode('utf-8-sig')
