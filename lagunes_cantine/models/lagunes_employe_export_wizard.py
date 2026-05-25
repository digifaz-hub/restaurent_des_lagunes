# -*- coding: utf-8 -*-

import io
import base64
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError


class LagunesEmployeExportWizard(models.TransientModel):
    """Wizard d'export de la liste des employés d'une entreprise (Excel + PDF)."""

    _name = 'lagunes.employe.export.wizard'
    _description = "Export des employés d'une entreprise"

    entreprise_id = fields.Many2one(
        'res.partner',
        string='Entreprise',
        domain=[('is_cantine_client', '=', True)],
        required=True,
        help="Laissez vide pour exporter tous les employés.",
    )

    actifs_seulement = fields.Boolean(
        string='Employés actifs seulement',
        default=True,
    )

    nb_employes = fields.Integer(
        string='Nombre d\'employés',
        compute='_compute_preview',
    )

    @api.depends('entreprise_id', 'actifs_seulement')
    def _compute_preview(self):
        for wizard in self:
            domain = wizard._build_domain()
            wizard.nb_employes = self.env['lagunes.employe'].search_count(domain)

    def _build_domain(self):
        domain = []
        if self.entreprise_id:
            domain.append(('entreprise_id', '=', self.entreprise_id.id))
        if self.actifs_seulement:
            domain.append(('active', '=', True))
        return domain

    def _get_employes(self):
        return self.env['lagunes.employe'].search(
            self._build_domain(),
            order='entreprise_id, nom, prenom'
        )

    # ──────────────────────────────────────────────────────────────
    #  EXPORT EXCEL
    # ──────────────────────────────────────────────────────────────

    def action_export_excel(self):
        self.ensure_one()
        employes = self._get_employes()
        if not employes:
            raise ValidationError(_('Aucun employé trouvé pour les critères sélectionnés.'))

        excel_data = self._generate_excel(employes)

        ent_slug = (self.entreprise_id.name or 'tous').replace(' ', '_')[:20]
        filename = f"employes_{ent_slug}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(excel_data),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def _generate_excel(self, employes):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ValidationError(
                _('La librairie openpyxl est requise pour l\'export Excel.\n'
                  'Installez-la avec : pip install openpyxl')
            )

        COLOR_BLUE  = "16166D"
        COLOR_LIGHT = "E0E1F0"
        COLOR_ALT   = "F7F7FC"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employés"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color=COLOR_BLUE, end_color=COLOR_BLUE, fill_type="solid")
        thin = Border(
            left=Side(style='thin', color='BBBBBB'),
            right=Side(style='thin', color='BBBBBB'),
            top=Side(style='thin', color='BBBBBB'),
            bottom=Side(style='thin', color='BBBBBB'),
        )

        # ── Titre ────────────────────────────────────────────────────────
        ws.merge_cells('A1:G1')
        ws['A1'].value = "Restaurant des Lagunes — Liste des Employés"
        ws['A1'].font = Font(bold=True, size=15, color=COLOR_BLUE)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        ws.merge_cells('A2:G2')
        ent_name = self.entreprise_id.name if self.entreprise_id else "Toutes les entreprises"
        ws['A2'].value = f"Entreprise : {ent_name}"
        ws['A2'].font = Font(bold=True, size=11, color=COLOR_BLUE)
        ws['A2'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A3:G3')
        statut = "actifs" if self.actifs_seulement else "actifs et inactifs"
        ws['A3'].value = (
            f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')} "
            f"— {len(employes)} employé(s) {statut}"
        )
        ws['A3'].font = Font(italic=True, color="888888", size=9)
        ws['A3'].alignment = Alignment(horizontal="center")

        # ── En-têtes colonnes ────────────────────────────────────────────
        headers = [
            ('Entreprise',       25),
            ('Nom',              20),
            ('Prénom',           20),
            ('Email',            32),
            ('Date inscription', 18),
            ('Commandes/jour',   16),
            ('Statut',           12),
        ]
        ROW_H = 5
        for col_num, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=ROW_H, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
            ws.column_dimensions[get_column_letter(col_num)].width = width
        ws.row_dimensions[ROW_H].height = 28

        alt_fill = PatternFill(start_color=COLOR_ALT, end_color=COLOR_ALT, fill_type="solid")

        for i, emp in enumerate(employes):
            r = ROW_H + 1 + i
            fill = alt_fill if i % 2 == 0 else None
            row_data = [
                emp.entreprise_id.name if emp.entreprise_id else '',
                emp.nom,
                emp.prenom,
                emp.email,
                emp.date_inscription.strftime('%d/%m/%Y') if emp.date_inscription else '',
                emp.max_commandes_par_jour or 1,
                'Actif' if emp.active else 'Inactif',
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col_num, value=value)
                cell.border = thin
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill
                # Coloration statut
                if col_num == 7:
                    if value == 'Actif':
                        cell.font = Font(color="2E7D32", bold=True)
                    else:
                        cell.font = Font(color="B71C1C", bold=True)

        # ── Ligne total ──────────────────────────────────────────────────
        last_data_row = ROW_H + len(employes)
        total_row = last_data_row + 2
        total_fill = PatternFill(start_color=COLOR_LIGHT, end_color=COLOR_LIGHT, fill_type="solid")

        ws.merge_cells(f'A{total_row}:F{total_row}')
        ws[f'A{total_row}'].value = f"TOTAL : {len(employes)} employé(s)"
        ws[f'A{total_row}'].font = Font(bold=True, color=COLOR_BLUE, size=11)
        ws[f'A{total_row}'].fill = total_fill
        ws[f'A{total_row}'].alignment = Alignment(horizontal="right", vertical="center")
        ws[f'A{total_row}'].border = thin
        ws[f'G{total_row}'].fill = total_fill
        ws[f'G{total_row}'].border = thin

        ws.freeze_panes = f'A{ROW_H + 1}'

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ──────────────────────────────────────────────────────────────
    #  EXPORT PDF
    # ──────────────────────────────────────────────────────────────

    def action_export_pdf(self):
        self.ensure_one()
        employes = self._get_employes()
        if not employes:
            raise ValidationError(_('Aucun employé trouvé pour les critères sélectionnés.'))

        pdf_data = self._generate_pdf(employes)

        ent_slug = (self.entreprise_id.name or 'tous').replace(' ', '_')[:20]
        filename = f"employes_{ent_slug}_{datetime.now().strftime('%Y%m%d')}.pdf"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(pdf_data),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/pdf',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def _generate_pdf(self, employes):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

        BLUE       = colors.HexColor('#16166D')
        BLUE_LIGHT = colors.HexColor('#E0E1F0')
        BLUE_ALT   = colors.HexColor('#F7F7FC')
        GREEN      = colors.HexColor('#2E7D32')
        RED        = colors.HexColor('#B71C1C')
        GREY_LIGHT = colors.HexColor('#888888')

        output = io.BytesIO()

        # Orientation paysage pour avoir plus d'espace
        doc = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
            title=f"Employés — {self.entreprise_id.name if self.entreprise_id else 'Toutes entreprises'}",
        )

        styles = getSampleStyleSheet()

        cell_style = ParagraphStyle(
            'TD', fontSize=8, textColor=colors.black,
            fontName='Helvetica', alignment=TA_LEFT, leading=10,
        )
        header_style = ParagraphStyle(
            'TH', fontSize=8, textColor=colors.white,
            fontName='Helvetica-Bold', alignment=TA_CENTER, leading=10,
        )
        cell_center = ParagraphStyle(
            'TDC', fontSize=8, textColor=colors.black,
            fontName='Helvetica', alignment=TA_CENTER, leading=10,
        )
        status_actif = ParagraphStyle(
            'ActifStyle', fontSize=8, textColor=GREEN,
            fontName='Helvetica-Bold', alignment=TA_CENTER, leading=10,
        )
        status_inactif = ParagraphStyle(
            'InactifStyle', fontSize=8, textColor=RED,
            fontName='Helvetica-Bold', alignment=TA_CENTER, leading=10,
        )

        story = []

        # ── En-tête ───────────────────────────────────────────────────────
        header_data = [[
            Paragraph("RESTAURANT DES LAGUNES", ParagraphStyle(
                'HTitle', fontSize=16, textColor=colors.white,
                fontName='Helvetica-Bold',
            )),
            Paragraph("LISTE DES EMPLOYÉS", ParagraphStyle(
                'HRight', fontSize=10, textColor=colors.HexColor('#E0E1F0'),
                fontName='Helvetica', alignment=TA_RIGHT,
            )),
        ]]
        header_table = Table(header_data, colWidths=[160 * mm, 80 * mm])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BLUE),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (0, 0), 6 * mm),
            ('RIGHTPADDING', (-1, 0), (-1, 0), 4 * mm),
            ('TOPPADDING', (0, 0), (-1, -1), 4 * mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4 * mm),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 4 * mm))

        # ── Infos document ────────────────────────────────────────────────
        ent_name = self.entreprise_id.name if self.entreprise_id else "Toutes les entreprises"
        statut = "actifs" if self.actifs_seulement else "actifs et inactifs"
        info_data = [[
            Paragraph(f"<b>Entreprise :</b> {ent_name}", ParagraphStyle(
                'Info', fontSize=11, textColor=BLUE, fontName='Helvetica-Bold',
            )),
            Paragraph(
                f"<b>Date :</b> {datetime.now().strftime('%d/%m/%Y à %H:%M')} — "
                f"{len(employes)} employé(s) {statut}",
                ParagraphStyle('InfoR', fontSize=9, textColor=colors.HexColor('#444'),
                               fontName='Helvetica', alignment=TA_RIGHT)
            ),
        ]]
        info_table = Table(info_data, colWidths=[160 * mm, 80 * mm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BLUE_LIGHT),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (0, 0), 4 * mm),
            ('RIGHTPADDING', (-1, 0), (-1, 0), 4 * mm),
            ('TOPPADDING', (0, 0), (-1, -1), 3 * mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3 * mm),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 5 * mm))

        # ── Tableau des employés ──────────────────────────────────────────
        # Largeurs colonnes en mm (total ≈ 242mm en paysage A4 avec marges 15mm×2)
        col_widths = [55 * mm, 35 * mm, 35 * mm, 65 * mm, 28 * mm, 24 * mm]

        table_data = [[
            Paragraph("Entreprise",        header_style),
            Paragraph("Nom",               header_style),
            Paragraph("Prénom",            header_style),
            Paragraph("Email",             header_style),
            Paragraph("Inscription",       header_style),
            Paragraph("Statut",            header_style),
        ]]

        for i, emp in enumerate(employes):
            statut_para = Paragraph(
                'Actif' if emp.active else 'Inactif',
                status_actif if emp.active else status_inactif
            )
            row = [
                Paragraph(emp.entreprise_id.name if emp.entreprise_id else '—', cell_style),
                Paragraph(emp.nom or '—',    cell_style),
                Paragraph(emp.prenom or '—', cell_style),
                Paragraph(emp.email or '—',  cell_style),
                Paragraph(
                    emp.date_inscription.strftime('%d/%m/%Y') if emp.date_inscription else '—',
                    cell_center
                ),
                statut_para,
            ]
            table_data.append(row)

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BLUE),
            ('ROWBACKGROUND', (0, 1), (-1, -1), [BLUE_ALT, colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BBBBBB')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 4 * mm))

        # ── Ligne total ────────────────────────────────────────────────────
        total_data = [[
            Paragraph(
                f"TOTAL : {len(employes)} employé(s)",
                ParagraphStyle('TotL', fontSize=10, textColor=BLUE,
                               fontName='Helvetica-Bold', alignment=TA_RIGHT)
            ),
        ]]
        total_tbl = Table(total_data, colWidths=[sum(col_widths)])
        total_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BLUE_LIGHT),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3 * mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3 * mm),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3 * mm),
            ('LINEABOVE', (0, 0), (-1, 0), 1.5, BLUE),
        ]))
        story.append(total_tbl)

        story.append(Spacer(1, 6 * mm))
        story.append(HRFlowable(width="100%", thickness=0.5, color=BLUE_LIGHT))

        doc.build(story)
        return output.getvalue()