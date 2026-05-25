# -*- coding: utf-8 -*-

import io
import base64
from datetime import date
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError


class LagunesExportWizard(models.TransientModel):
    _name = 'lagunes.export.wizard'
    _description = 'Assistant d\'export des commandes'

    date_debut = fields.Date(
        string='Date de début',
        default=lambda self: date.today().replace(day=1),
        required=True,
    )

    date_fin = fields.Date(
        string='Date de fin',
        default=fields.Date.today,
        required=True,
    )

    # FIX : entreprise obligatoire
    entreprise_id = fields.Many2one(
        'res.partner',
        string='Entreprise',
        domain=[('is_cantine_client', '=', True)],
        required=True,
        help='Sélectionnez l\'entreprise à exporter.',
    )

    nb_commandes = fields.Integer(
        string='Nombre de commandes',
        compute='_compute_preview',
    )

    montant_total = fields.Float(
        string='Montant total',
        compute='_compute_preview',
    )

    # ------------------------------------------------------------------ #
    #  COMPUTE                                                             #
    # ------------------------------------------------------------------ #

    @api.depends('date_debut', 'date_fin', 'entreprise_id')
    def _compute_preview(self):
        for wizard in self:
            domain = wizard._build_domain()
            commandes = self.env['lagunes.commande'].search(domain)
            wizard.nb_commandes = len(commandes)
            wizard.montant_total = sum(commandes.mapped('prix_total'))

    # ------------------------------------------------------------------ #
    #  HELPERS                                                             #
    # ------------------------------------------------------------------ #

    def _build_domain(self):
        domain = []
        if self.date_debut:
            domain.append(('date', '>=', self.date_debut))
        if self.date_fin:
            domain.append(('date', '<=', self.date_fin))
        if self.entreprise_id:
            domain.append(('entreprise_id', '=', self.entreprise_id.id))
        return domain

    def _get_plats_from_commande(self, cmd):
        """Retourne (entree, plat_principal, dessert) pour une commande.

        Priorité à l'architecture moderne `line_ids` : groupe les lignes par
        type de plat (entrée / résistance / dessert) via le nom du type.
        Si la commande n'a pas de lignes (ancien format), fallback sur les
        champs legacy `entree_plat_id` / `resistance_plat_id` / `dessert_plat_id`.

        Retour : tuple de 3 strings (jamais None ; '—' si absent).
        """
        DASH = '—'

        if cmd.line_ids:
            entrees, plats, desserts = [], [], []
            for line in cmd.line_ids:
                type_name = (line.plat_type_id.name or '').lower()
                plat_label = (
                    line.plat_id.display_name_website
                    or line.plat_id.name
                    or ''
                )
                if not plat_label:
                    continue
                if 'entr' in type_name:
                    entrees.append(plat_label)
                elif 'sist' in type_name or 'principal' in type_name:
                    plats.append(plat_label)
                elif 'dessert' in type_name:
                    desserts.append(plat_label)
                else:
                    # Type inconnu : on l'ajoute aux plats principaux par défaut
                    plats.append(plat_label)
            return (
                ' + '.join(entrees) if entrees else DASH,
                ' + '.join(plats) if plats else DASH,
                ' + '.join(desserts) if desserts else DASH,
            )

        # Fallback architecture legacy
        entree = (
            cmd.entree_plat_id.display_name_website
            or cmd.entree_plat_id.name
        ) if cmd.entree_plat_id else DASH
        plat = (
            cmd.resistance_plat_id.display_name_website
            or cmd.resistance_plat_id.name
        ) if cmd.resistance_plat_id else DASH
        dessert = (
            cmd.dessert_plat_id.display_name_website
            or cmd.dessert_plat_id.name
        ) if cmd.dessert_plat_id else DASH
        return entree, plat, dessert

    def _get_period_label(self):
        """Retourne le libellé de période : date unique ou intervalle."""
        if self.date_debut and self.date_fin and self.date_debut == self.date_fin:
            return self.date_debut.strftime('%d/%m/%Y')
        parts = []
        if self.date_debut:
            parts.append(f"Du {self.date_debut.strftime('%d/%m/%Y')}")
        if self.date_fin:
            parts.append(f"au {self.date_fin.strftime('%d/%m/%Y')}")
        return " ".join(parts) if parts else "Toutes les commandes"

    # ------------------------------------------------------------------ #
    #  ACTIONS                                                             #
    # ------------------------------------------------------------------ #

    def action_export_excel(self):
        self.ensure_one()

        if self.date_fin < self.date_debut:
            raise ValidationError(_('La date de fin doit être postérieure à la date de début.'))

        domain = self._build_domain()
        commandes = self.env['lagunes.commande'].search(
            domain,
            order='date asc, create_date asc'
        )

        if not commandes:
            raise ValidationError(_('Aucune commande trouvée pour les critères sélectionnés.'))

        excel_data = self._generate_excel(commandes)

        date_debut_str = self.date_debut.strftime('%Y%m%d') if self.date_debut else 'debut'
        date_fin_str = self.date_fin.strftime('%Y%m%d') if self.date_fin else 'fin'
        ent_slug = (self.entreprise_id.name or 'entreprise').replace(' ', '_')[:20]
        filename = f"commandes_{ent_slug}_{date_debut_str}_{date_fin_str}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(excel_data),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    def action_export_pdf(self):
        self.ensure_one()

        if self.date_fin < self.date_debut:
            raise ValidationError(_('La date de fin doit être postérieure à la date de début.'))

        domain = self._build_domain()
        commandes = self.env['lagunes.commande'].search(
            domain,
            order='date asc, create_date asc'
        )

        if not commandes:
            raise ValidationError(_('Aucune commande trouvée pour les critères sélectionnés.'))

        pdf_data = self._generate_pdf(commandes)

        date_debut_str = self.date_debut.strftime('%Y%m%d') if self.date_debut else 'debut'
        date_fin_str = self.date_fin.strftime('%Y%m%d') if self.date_fin else 'fin'
        ent_slug = (self.entreprise_id.name or 'entreprise').replace(' ', '_')[:20]
        filename = f"commandes_{ent_slug}_{date_debut_str}_{date_fin_str}.pdf"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(pdf_data),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/pdf',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'new',
        }

    # ------------------------------------------------------------------ #
    #  GÉNÉRATION EXCEL                                                    #
    # ------------------------------------------------------------------ #

    def _generate_excel(self, commandes):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ValidationError(
                _('La librairie openpyxl est requise pour l\'export Excel.\n'
                  'Installez-la avec : pip install openpyxl')
            )

        from datetime import datetime

        wb = openpyxl.Workbook()

        # ── Feuille unique : Détail des commandes ─────────────────────
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(title="Commandes")
        ws.title = "Commandes"

        COLOR_BLUE  = "16166D"
        COLOR_LIGHT = "E0E1F0"
        COLOR_ALT   = "F7F7FC"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill(start_color=COLOR_BLUE, end_color=COLOR_BLUE, fill_type="solid")
        thin = Border(
            left=Side(style='thin', color='BBBBBB'),
            right=Side(style='thin', color='BBBBBB'),
            top=Side(style='thin', color='BBBBBB'),
            bottom=Side(style='thin', color='BBBBBB'),
        )

        # Titre
        ws.merge_cells('A1:G1')
        ws['A1'].value = "Restaurant des Lagunes — Commandes cantine"
        ws['A1'].font = Font(bold=True, size=15, color=COLOR_BLUE)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # Entreprise
        ws.merge_cells('A2:G2')
        ws['A2'].value = f"Entreprise : {self.entreprise_id.name}"
        ws['A2'].font = Font(bold=True, size=11, color=COLOR_BLUE)
        ws['A2'].alignment = Alignment(horizontal="center")

        # Période — date unique ou intervalle
        period_text = self._get_period_label()
        ws.merge_cells('A3:G3')
        ws['A3'].value = period_text
        ws['A3'].font = Font(italic=True, color="444444", size=10)
        ws['A3'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A4:G4')
        ws['A4'].value = (
            f"Exporté le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
            f" — {len(commandes)} commande(s)"
        )
        ws['A4'].font = Font(italic=True, color="888888", size=9)
        ws['A4'].alignment = Alignment(horizontal="center")

        # En-têtes colonnes — 7 colonnes (A à G) avec détail des plats
        headers = [
            ('Date', 12),
            ('Employé', 20),
            ('Entrée', 18),
            ('Plat principal', 18),
            ('Dessert', 18),
            ('Nb couverts', 10),
            ('Prix total\n(FCFA)', 14),
        ]
        ROW_H = 6
        for col_num, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=ROW_H, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin
            ws.column_dimensions[get_column_letter(col_num)].width = width
        ws.row_dimensions[ROW_H].height = 32

        alt_fill = PatternFill(start_color=COLOR_ALT, end_color=COLOR_ALT, fill_type="solid")

        for i, cmd in enumerate(commandes):
            r = ROW_H + 1 + i
            fill = alt_fill if i % 2 == 0 else None
            # Extraction des plats : priorité à `line_ids`, fallback legacy
            entree, plat_principal, dessert = self._get_plats_from_commande(cmd)

            row_data = [
                cmd.date.strftime('%d/%m/%Y') if cmd.date else '',
                cmd.employee_name or '',
                entree,
                plat_principal,
                dessert,
                cmd.quantity,
                cmd.prix_total,
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=col_num, value=value)
                cell.border = thin
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill
                if col_num == 7:  # Prix total
                    cell.number_format = '#,##0'

        # Ligne total
        last_data_row = ROW_H + len(commandes)
        total_row = last_data_row + 2
        total_fill = PatternFill(start_color=COLOR_LIGHT, end_color=COLOR_LIGHT, fill_type="solid")

        ws.merge_cells(f'A{total_row}:F{total_row}')
        ws[f'A{total_row}'].value = f"TOTAL — {len(commandes)} commande(s)"
        ws[f'A{total_row}'].font = Font(bold=True, color=COLOR_BLUE, size=11)
        ws[f'A{total_row}'].fill = total_fill
        ws[f'A{total_row}'].alignment = Alignment(horizontal="right", vertical="center")
        ws[f'A{total_row}'].border = thin

        total_val = ws.cell(
            row=total_row,
            column=7,
            value=sum(commandes.mapped('prix_total'))
        )
        total_val.font = Font(bold=True, size=11, color=COLOR_BLUE)
        total_val.number_format = '#,##0'
        total_val.fill = total_fill
        total_val.border = thin
        total_val.alignment = Alignment(horizontal="right", vertical="center")

        ws.freeze_panes = f'A{ROW_H + 1}'

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ------------------------------------------------------------------ #
    #  GÉNÉRATION PDF                                                      #
    # ------------------------------------------------------------------ #

    def _generate_pdf(self, commandes):
        """Génère un PDF professionnel des commandes via ReportLab."""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from datetime import datetime

        # ── Couleurs ───────────────────────────────────────────────────
        BLUE       = colors.HexColor('#16166D')
        BLUE_LIGHT = colors.HexColor('#E0E1F0')
        BLUE_ALT   = colors.HexColor('#F7F7FC')
        ORANGE     = colors.HexColor('#F5A821')
        GREY_TEXT  = colors.HexColor('#444444')
        GREY_LIGHT = colors.HexColor('#888888')

        output = io.BytesIO()

        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=15 * mm,
            leftMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
            title=f"Commandes — {self.entreprise_id.name}",
        )

        styles = getSampleStyleSheet()

        # Styles personnalisés
        s_title = ParagraphStyle(
            'LagunesTitle',
            parent=styles['Normal'],
            fontSize=18,
            leading=22,
            textColor=BLUE,
            fontName='Helvetica-Bold',
            alignment=TA_LEFT,
        )
        s_subtitle = ParagraphStyle(
            'LagunesSubtitle',
            parent=styles['Normal'],
            fontSize=11,
            textColor=GREY_TEXT,
            fontName='Helvetica-Bold',
            alignment=TA_LEFT,
        )
        s_meta = ParagraphStyle(
            'LaguesMeta',
            parent=styles['Normal'],
            fontSize=8,
            textColor=GREY_LIGHT,
            fontName='Helvetica',
            alignment=TA_LEFT,
        )
        s_total_label = ParagraphStyle(
            'LaguesTotalLabel',
            parent=styles['Normal'],
            fontSize=11,
            textColor=BLUE,
            fontName='Helvetica-Bold',
            alignment=TA_RIGHT,
        )
        s_total_value = ParagraphStyle(
            'LaguesTotalValue',
            parent=styles['Normal'],
            fontSize=11,
            textColor=BLUE,
            fontName='Helvetica-Bold',
            alignment=TA_RIGHT,
        )

        story = []

        # ── En-tête ────────────────────────────────────────────────────
        # Bandeau bleu supérieur simulé avec un tableau
        header_data = [[
            Paragraph("RESTAURANT DES LAGUNES", ParagraphStyle(
                'HeaderTitle',
                fontSize=16,
                textColor=colors.white,
                fontName='Helvetica-Bold',
            )),
            Paragraph("COMMANDES CANTINE", ParagraphStyle(
                'HeaderRight',
                fontSize=10,
                textColor=colors.HexColor('#E0E1F0'),
                fontName='Helvetica',
                alignment=TA_RIGHT,
            )),
        ]]
        header_table = Table(header_data, colWidths=[110 * mm, 65 * mm])
        header_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BLUE),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (0, 0), 6 * mm),
            ('RIGHTPADDING', (-1, 0), (-1, 0), 4 * mm),
            ('TOPPADDING', (0, 0), (-1, -1), 4 * mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4 * mm),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 4 * mm))

        # ── Infos document ─────────────────────────────────────────────
        period_text = self._get_period_label()
        info_data = [[
            Paragraph(f"<b>Entreprise :</b> {self.entreprise_id.name}", ParagraphStyle(
                'Info',
                fontSize=11,
                textColor=BLUE,
                fontName='Helvetica-Bold',
            )),
            Paragraph(
                f"<b>Période :</b> {period_text}",
                ParagraphStyle('InfoR', fontSize=10, textColor=GREY_TEXT,
                               fontName='Helvetica', alignment=TA_RIGHT)
            ),
        ]]
        info_table = Table(info_data, colWidths=[110 * mm, 65 * mm])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BLUE_LIGHT),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (0, 0), 4 * mm),
            ('RIGHTPADDING', (-1, 0), (-1, 0), 4 * mm),
            ('TOPPADDING', (0, 0), (-1, -1), 3 * mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3 * mm),
        ]))
        story.append(info_table)

        story.append(Spacer(1, 2 * mm))

        meta_text = (
            f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — "
            f"{len(commandes)} commande(s)"
        )
        story.append(Paragraph(meta_text, s_meta))
        story.append(Spacer(1, 5 * mm))

        # ── Tableau des commandes ──────────────────────────────────────
        # 7 colonnes : Date, Employé, Entrée, Plat, Dessert, Qté, Total
        col_widths = [20 * mm, 28 * mm, 22 * mm, 22 * mm, 22 * mm, 14 * mm, 16 * mm]

        header_style = ParagraphStyle(
            'TH',
            fontSize=8,
            textColor=colors.white,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            leading=10,
        )
        cell_style = ParagraphStyle(
            'TD',
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica',
            alignment=TA_LEFT,
            leading=10,
        )
        cell_center = ParagraphStyle(
            'TDC',
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica',
            alignment=TA_CENTER,
            leading=10,
        )
        cell_right = ParagraphStyle(
            'TDR',
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica',
            alignment=TA_RIGHT,
            leading=10,
        )

        def fmt_number(val):
            """Formatage nombre avec espace millier."""
            try:
                return f"{int(val):,}".replace(',', ' ')
            except Exception:
                return str(val)

        table_data = [[
            Paragraph("Date", header_style),
            Paragraph("Employé", header_style),
            Paragraph("Entrée", header_style),
            Paragraph("Plat", header_style),
            Paragraph("Dessert", header_style),
            Paragraph("Nb couverts", header_style),
            Paragraph("Total", header_style),
        ]]

        for i, cmd in enumerate(commandes):
            # Extraction des plats : priorité à `line_ids`, fallback legacy
            entree, plat_principal, dessert = self._get_plats_from_commande(cmd)

            row = [
                Paragraph(cmd.date.strftime('%d/%m/%Y') if cmd.date else '', cell_center),
                Paragraph(cmd.employee_name or '—', cell_style),
                Paragraph(entree, cell_style),
                Paragraph(plat_principal, cell_style),
                Paragraph(dessert, cell_style),
                Paragraph(str(cmd.quantity), cell_center),
                Paragraph(fmt_number(cmd.prix_total), cell_right),
            ]
            table_data.append(row)

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

        tbl_style = [
            # En-tête
            ('BACKGROUND', (0, 0), (-1, 0), BLUE),
            ('ROWBACKGROUND', (0, 1), (-1, -1), [BLUE_ALT, colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#BBBBBB')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]
        tbl.setStyle(TableStyle(tbl_style))

        story.append(tbl)
        story.append(Spacer(1, 4 * mm))

        # ── Ligne total ────────────────────────────────────────────────
        grand_total = sum(commandes.mapped('prix_total'))
        total_data = [[
            Paragraph(
                f"TOTAL — {len(commandes)} commande(s)",
                ParagraphStyle('TotL', fontSize=10, textColor=BLUE,
                               fontName='Helvetica-Bold', alignment=TA_RIGHT)
            ),
            Paragraph(
                f"{fmt_number(grand_total)} FCFA",
                ParagraphStyle('TotV', fontSize=11, textColor=BLUE,
                               fontName='Helvetica-Bold', alignment=TA_RIGHT)
            ),
        ]]
        total_tbl = Table(total_data, colWidths=[130 * mm, 30 * mm])
        total_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BLUE_LIGHT),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3 * mm),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3 * mm),
            ('RIGHTPADDING', (-1, 0), (-1, 0), 3 * mm),
            ('LEFTPADDING', (0, 0), (0, 0), 3 * mm),
            ('LINEABOVE', (0, 0), (-1, 0), 1.5, BLUE),
        ]))
        story.append(total_tbl)

        story.append(Spacer(1, 6 * mm))

        # ── Pied de page ───────────────────────────────────────────────
        story.append(HRFlowable(width="100%", thickness=0.5, color=BLUE_LIGHT))
        story.append(Spacer(1, 2 * mm))
        # story.append(Paragraph(
        #     "Restaurant des Lagunes — Document confidentiel à usage interne",
        #     ParagraphStyle('Footer', fontSize=7, textColor=GREY_LIGHT,
        #                    fontName='Helvetica', alignment=TA_CENTER)
        # ))

        doc.build(story)
        return output.getvalue()